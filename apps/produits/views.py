from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator
from django.db.models import Q, F, Sum
from django.http import JsonResponse, HttpResponse
from .models import Produit, Categorie, MouvementStock, HistoriquePrix
from .forms import ProduitForm, CategorieForm, AjustementStockForm
from apps.core.sanitizers import sanitize_search_query

@login_required
def liste(request):
    q   = sanitize_search_query(request.GET.get('q',''))
    cat = request.GET.get('categorie','')
    qs  = Produit.objects.filter(supprime=False, actif=True).select_related('categorie')
    if q:   qs = qs.filter(Q(nom__icontains=q)|Q(reference__icontains=q)|Q(code_barre__icontains=q))
    if cat: qs = qs.filter(categorie_id=cat)
    categories = Categorie.objects.all().order_by('nom')
    alertes    = Produit.objects.filter(quantite_stock__lte=F('seuil_alerte'), actif=True, supprime=False).count()
    page = Paginator(qs.order_by('nom'), 25).get_page(request.GET.get('page'))
    return render(request, 'produits/liste.html', {'produits': page, 'q': q, 'categories': categories, 'cat_active': cat, 'alertes_count': alertes})

@login_required
def detail(request, pk):
    produit  = get_object_or_404(Produit, pk=pk, supprime=False)
    mouvements = MouvementStock.objects.filter(produit=produit).order_by('-cree_le')[:20]
    historique_prix = HistoriquePrix.objects.filter(produit=produit).select_related('modifie_par').order_by('-cree_le')[:10]
    return render(request, 'produits/detail.html', {'produit': produit, 'mouvements': mouvements, 'historique_prix': historique_prix})

@login_required
def create(request):
    form = ProduitForm(request.POST or None, request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Produit cree.")
        return redirect('produits:liste')
    return render(request, 'produits/form.html', {'form': form, 'titre': 'Nouveau produit'})

@login_required
def edit(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    ancien_achat = produit.prix_achat
    ancien_vente = produit.prix_vente
    form = ProduitForm(request.POST or None, request.FILES or None, instance=produit)
    if form.is_valid():
        p = form.save()
        if p.prix_achat != ancien_achat or p.prix_vente != ancien_vente:
            HistoriquePrix.objects.create(
                produit=p, ancien_prix_achat=ancien_achat, nouveau_prix_achat=p.prix_achat,
                ancien_prix_vente=ancien_vente, nouveau_prix_vente=p.prix_vente, modifie_par=request.user
            )
        messages.success(request, "Produit mis a jour.")
        return redirect('produits:detail', pk=p.pk)
    return render(request, 'produits/form.html', {'form': form, 'titre': f'Modifier — {produit.nom}', 'produit': produit})

@login_required
def delete(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    if request.method == 'POST':
        produit.supprimer()
        messages.success(request, f"Produit {produit.nom} supprime.")
        return redirect('produits:liste')
    return render(request, 'produits/confirm_delete.html', {'produit': produit})

@login_required
def ajuster_stock(request, pk):
    produit = get_object_or_404(Produit, pk=pk)
    form    = AjustementStockForm(request.POST or None)
    if form.is_valid():
        from decimal import Decimal
        cd        = form.cleaned_data
        qte       = cd['quantite']
        type_mv   = cd['type_mouvement']
        stock_av  = produit.quantite_stock
        if type_mv == 'entree':
            produit.quantite_stock += qte
        else:
            produit.quantite_stock = max(Decimal('0'), produit.quantite_stock - qte)
        produit.save()
        MouvementStock.objects.create(
            produit=produit, type_mouvement=type_mv, quantite=qte,
            stock_avant=stock_av, stock_apres=produit.quantite_stock,
            raison=cd.get('raison',''), utilisateur=request.user,
        )
        messages.success(request, f"Stock ajuste : {stock_av} → {produit.quantite_stock} {produit.unite}")
        return redirect('produits:detail', pk=pk)
    return render(request, 'produits/ajuster_stock.html', {'form': form, 'produit': produit})

@login_required
def alertes(request):
    produits = Produit.objects.filter(quantite_stock__lte=F('seuil_alerte'), actif=True, supprime=False).select_related('categorie','fournisseur').order_by('quantite_stock')
    return render(request, 'produits/alertes.html', {'produits': produits})

@login_required
def categories(request):
    cats = Categorie.objects.annotate(nb=Q()).order_by('nom')
    return render(request, 'produits/categories.html', {'categories': Categorie.objects.all().order_by('nom')})

@login_required
def categorie_create(request):
    form = CategorieForm(request.POST or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Categorie creee.")
        return redirect('produits:categories')
    return render(request, 'produits/categorie_form.html', {'form': form, 'titre': 'Nouvelle categorie'})

@login_required
def categorie_edit(request, pk):
    cat  = get_object_or_404(Categorie, pk=pk)
    form = CategorieForm(request.POST or None, instance=cat)
    if form.is_valid():
        form.save()
        messages.success(request, "Categorie mise a jour.")
        return redirect('produits:categories')
    return render(request, 'produits/categorie_form.html', {'form': form, 'titre': f'Modifier — {cat.nom}', 'categorie': cat})

@login_required
def search_ajax(request):
    q  = sanitize_search_query(request.GET.get('q',''))
    qs = Produit.objects.filter(Q(nom__icontains=q)|Q(reference__icontains=q), supprime=False, actif=True)[:10]
    data = [{
        'id': p.pk,
        'reference': p.reference,
        'nom': p.nom,
        'prix_vente_ht': float(p.prix_vente),
        'prix_achat_ht': float(p.prix_achat),
        'tva': float(p.tva),
        'stock': float(p.quantite_stock),
        'unite': p.get_unite_display(),
    } for p in qs]
    return JsonResponse({'results': data})

@login_required
def tarif_preview(request):
    cat_id = request.GET.get('categorie','')
    qs     = Produit.objects.filter(actif=True, supprime=False, peut_vendre=True).select_related('categorie').order_by('categorie__nom','nom')
    if cat_id: qs = qs.filter(categorie_id=cat_id)
    return render(request, 'produits/tarif_preview.html', {'produits': qs, 'categories': Categorie.objects.all().order_by('nom'), 'categorie_id': cat_id})

@login_required
def tarif_pdf(request):
    cat_id = request.GET.get('categorie','')
    qs     = Produit.objects.filter(actif=True, supprime=False, peut_vendre=True).select_related('categorie').order_by('categorie__nom','nom')
    if cat_id: qs = qs.filter(categorie_id=cat_id)
    from apps.rapports.pdf_generator import generer_pdf_tarif
    return generer_pdf_tarif(qs)

@login_required
def export_stock_excel(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire Stock"
    headers = ['Reference','Designation','Categorie','Unite','Stock','Seuil','PU Achat','PU Vente','Valeur Stock','Marge %','Statut']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
    for ri, p in enumerate(Produit.objects.filter(actif=True, supprime=False).select_related('categorie').order_by('categorie__nom','nom'), 2):
        statut = 'Rupture' if p.quantite_stock == 0 else ('Alerte' if p.en_alerte else 'OK')
        ws.append([p.reference, p.nom, p.categorie.nom if p.categorie else '', p.get_unite_display(),
                   float(p.quantite_stock), float(p.seuil_alerte), float(p.prix_achat), float(p.prix_vente),
                   round(float(p.quantite_stock)*float(p.prix_achat),2), round(float(p.marge_brute),1), statut])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="inventaire_stock.xlsx"'
    wb.save(response)
    return response

@login_required
def import_excel(request):
    return render(request, 'produits/import_excel.html', {})

@login_required
def import_preview(request):
    if request.method == 'POST' and request.POST.get('confirmed') == '1':
        import json
        from decimal import Decimal
        lignes = json.loads(request.POST.get('lignes_json','[]'))
        created = updated = 0
        for row in lignes:
            ref  = str(row.get('reference','')).strip()
            nom  = str(row.get('designation','')).strip()
            if not ref or not nom: continue
            cat_nom = str(row.get('categorie','')).strip()
            categorie = None
            if cat_nom:
                categorie, _ = Categorie.objects.get_or_create(nom=cat_nom)
            defaults = {
                'nom': nom, 'categorie': categorie,
                'prix_achat': Decimal(str(row.get('prix_achat',0) or 0)),
                'prix_vente': Decimal(str(row.get('prix_vente',0) or 0)),
                'tva':        Decimal(str(row.get('tva',19) or 19)),
                'quantite_stock': Decimal(str(row.get('stock',0) or 0)),
                'seuil_alerte':   Decimal(str(row.get('seuil_alerte',5) or 5)),
            }
            _, is_new = Produit.objects.update_or_create(reference=ref, defaults=defaults)
            if is_new: created += 1
            else: updated += 1
        messages.success(request, f"{created} cree(s), {updated} mis a jour.")
        return redirect('produits:liste')
    if 'fichier' not in request.FILES:
        messages.error(request, "Aucun fichier fourni.")
        return redirect('produits:import_excel')

    # Verify uploaded file extension before processing
    f = request.FILES['fichier']
    ALLOWED_EXT = ('.xlsx', '.xlsm', '.xltx', '.xls')
    if not any(f.name.lower().endswith(ext) for ext in ALLOWED_EXT):
        messages.error(request, "Type de fichier non autorise. Formats acceptes: xlsx, xlsm, xltx, xls.")
        return redirect('produits:import_excel')
    try:
        import openpyxl, json
        wb  = openpyxl.load_workbook(request.FILES['fichier'], read_only=True, data_only=True)
        ws  = wb.active
        rows= ws.iter_rows(values_only=True)
        hdr = [str(h or '').strip().lower().replace(' ','_') for h in (next(rows, []) or [])]
        col = lambda key, aliases: next((hdr.index(a) for a in aliases if a in hdr), None)
        ci  = {
            'reference':   col('reference',   ['reference','ref','code','sku']),
            'designation': col('designation', ['designation','nom','produit','name']),
            'categorie':   col('categorie',   ['categorie','category','cat']),
            'prix_achat':  col('prix_achat',  ['prix_achat','achat','cout']),
            'prix_vente':  col('prix_vente',  ['prix_vente','vente','prix','price']),
            'tva':         col('tva',         ['tva','tax','taxe']),
            'stock':       col('stock',       ['stock','quantite','qty']),
            'seuil_alerte':col('seuil_alerte',['seuil_alerte','seuil','alerte']),
        }
        def gcell(row, key):
            idx = ci.get(key)
            return row[idx] if idx is not None and idx < len(row) else None
        def sfloat(v, d=0):
            try: return float(str(v).replace(',','.').strip())
            except: return d
        lignes = []
        for rn, row in enumerate(rows, 2):
            ref = str(gcell(row,'reference') or '').strip()
            if not ref: continue
            exists = Produit.objects.filter(reference=ref).exists()
            lignes.append({
                'row_num': rn, 'reference': ref,
                'designation': str(gcell(row,'designation') or '').strip(),
                'categorie':   str(gcell(row,'categorie') or '').strip(),
                'prix_achat':  sfloat(gcell(row,'prix_achat')),
                'prix_vente':  sfloat(gcell(row,'prix_vente')),
                'tva':         sfloat(gcell(row,'tva'), 19),
                'stock':       sfloat(gcell(row,'stock')),
                'seuil_alerte':sfloat(gcell(row,'seuil_alerte'), 5),
                'action':      'mise a jour' if exists else 'creation',
            })
        return render(request, 'produits/import_preview.html', {
            'lignes': lignes, 'lignes_json': json.dumps(lignes),
            'nb_creation': sum(1 for l in lignes if l['action']=='creation'),
            'nb_update':   sum(1 for l in lignes if l['action']=='mise a jour'),
        })
    except Exception as e:
        messages.error(request, f"Erreur : {str(e)[:100]}")
        return redirect('produits:import_excel')
